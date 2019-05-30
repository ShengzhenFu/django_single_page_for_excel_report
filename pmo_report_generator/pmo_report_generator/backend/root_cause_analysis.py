import os
import pandas as pd
import openpyxl


"""""""""""""""""""""""""""
author: Shengzhen.Fu
purpose: load RCA breakdown data from P1P2incidents_All file then update Root Cause Analysis in kun report
version: 1.0.0
date: 2/Feb/2019
"""""""""""""""""""""""""""

class sheet4_update:
    def read_rca_data(srcfile):
        """
        get rca and category data from source and save to dictionary
        :param srcfile:
        :return:
        """
        print('start to read open action data from action item list file')
        if os.path.exists(srcfile):
            # A key, R root cause breakdown
            df = pd.read_excel(srcfile, sheet_name='ALL', usecols="A, R")
            px = df[(df['root cause breakdown']).notnull()].groupby(['root cause breakdown']).size().reset_index(
                name='Count')
            py = px.sort_values(by=['Count'], ascending=False)
            res_dict = py.to_dict('list')
            # print(res_dict)
            return res_dict
            print('complete read open action data from action item list file')
        else:
            print('file %s not found %s', srcfile)

    def update_rca_data(srcfile, tarfile):
        """
        count rca by category and update to excel
        :param srcfile:
        :param tarfile:
        :return:
        """
        try:
            print('loading kun report file')
            wb = openpyxl.load_workbook(filename=tarfile)
        except Exception as e:
            print('file of kun report is opened by another program', e)
        sheet = wb['Root Cause Analysis']  # sheet name is Root Cause Analysis
        row_id = 5
        res_dict = sheet4_update.read_rca_data(srcfile)
        j = 0
        if len(res_dict.values()) > 0:
            print('begin to update RCA data to kun report')
            for x, y in enumerate(res_dict.values()):
                col = format(y).replace('[', '').replace(']', '').replace('\'', '')
                list_col = col.split(", ")
                for i in list_col:
                    if not i.isdigit():
                        sheet.cell(row=row_id, column=1).value = i
                        row_id += 1
                        j += 1
                    if i.isdigit():
                        sheet.cell(row=row_id - j, column=2).value = float(i)
                        row_id = row_id+1
                wb.save(tarfile)
            print('complete update open actionItems data to kun report')
        else:
            print('found no data in P1&P2 Incident 201701_201901_All, please check the data in file')

