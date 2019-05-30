import os
import pandas as pd
import statistics
import openpyxl
import datetime

"""""""""""""""""""""""""""
author: Shengzhen.Fu
purpose: load P1P2 response time data from P1P2incidents_All file then update response time in kun report
version: 1.0.0
date: 2/Feb/2019
"""""""""""""""""""""""""""


class sheet2_update:

    def get_last_day_previous_month():
        from datetime import date
        from dateutil.relativedelta import relativedelta
        today = date.today()
        return date(today.year, today.month, 1) - relativedelta(days=1)


    def get_respoinse_time_by_month(file, priority, month):
        """
        calculate response time in MEDIAN, by month and priority
        :param file:
        :param priority:
        :param month:
        :return:
        """
        # print('start to read response time from ')
        if priority in ['P1', 'P2']:
            if os.path.exists(file):
                # A key, C priority, M created, Q incident cause category, S top 6 product, X response hours, AB sz involved
                df = pd.read_excel(file, sheet_name='ALL', usecols="A,C,M,Q,S,X,AB")
                df.loc[:, ('yearMonth')] = df['created'].map(lambda x: 100 * x.year + x.month)
                px = df[(df['created'] >= '2018-01-01') & (df['top 6 product'] == 'Yes') & (df['sz involved'] == 'Yes') &
                        (df['priority']==priority) & (df['response hours'] > 0) & (df['yearMonth'].astype(str) == month)]
                if len(px) > 0:
                    """
                    calculate response time by MEDIAN method
                    """
                    median_resp_time = round(statistics.median(set(px['response hours'].tolist())), 2)
                    return median_resp_time
                else:
                    #print('no data found in priority', priority, 'on year month', month, 'from file', file)
                    return 0
            else:
                print('file %s not found' % file)
                return 'file not found'
        else:
            print('priority is not correct, make sure it is P1 or P2')
            return 'wrong priority value recieved'

    def get_all_month_list():
        """
        get month list from 2018-01 to last month
        :return:
        """
        start_date = "2018-01-01"  # input start date
        end_date = sheet2_update.get_last_day_previous_month()  # input end date
        month_list = [i.strftime("%Y-%m") for i in pd.date_range(start=start_date, end=end_date, freq='MS')]
        return month_list

    def get_all_month_response_time(priority, srcfile):
        """
        calculate response time (median) for all month by priority
        :param priority:
        :param srcfile:
        :return:
        """
        start_date = "2018-01-01"  # input start date
        # end_date = datetime.datetime.now()  # input end date
        end_date = sheet2_update.get_last_day_previous_month()
        month_list = [i.strftime("%Y%m") for i in pd.date_range(start=start_date, end=end_date, freq='MS')]
        p_list = {}
        r_list = []
        mon_list = []
        for i_month in month_list:
            month_obj = datetime.datetime.strptime(str(i_month), "%Y%m")  # convert month string to date object
            p = sheet2_update.get_respoinse_time_by_month(srcfile, priority, i_month)
            p_list.update({month_obj.strftime("%Y-%m"): p})
        for key, value in p_list.items():
            r_time = [value]
            mon = [key]
            r_list.append(r_time)
            mon_list.append(mon)
        return r_list

    def update_response_time_sheet(srcfile, tarfile):
        """
        update month, P1, P2 column to target file
        :param srcfile:
        :param file:
        :return:
        """
        # check if file exists
        if os.path.exists(tarfile):
            try:
                wb = openpyxl.load_workbook(filename=tarfile)
                sheet = wb['Response Time']  # sheet name is Response Time
                row_id = 4
                col_id = 1
                for x in sheet2_update.get_all_month_list():
                    sheet.cell(row=row_id, column=col_id).value = x
                    row_id += 1
                wb.save(tarfile)
                print('month has been updated to response sheet')
            except Exception as e:
                print('error occurred when update month list to response time sheet in Kun report', e)
            try:
                wb = openpyxl.load_workbook(filename=tarfile)
                sheet = wb['Response Time']
                row_id = 4
                col_id = 2
                for y in sheet2_update.get_all_month_response_time('P1', srcfile):
                    sheet.cell(row=row_id, column=col_id).value = float(str(y).replace('[', '').replace(']', ''))
                    row_id += 1
            except Exception as e:
                print('error occurred when load P1 data', e)
            try:
                wb.save(tarfile)
                print('P1 response time has been updated to response sheet')
            except Exception as e:
                print('error occurred when save data to Response Time sheet', e)
            try:
                wb = openpyxl.load_workbook(filename=tarfile)
                sheet = wb['Response Time']
                row_id = 4
                col_id = 3
                for y in sheet2_update.get_all_month_response_time('P2', srcfile):
                    sheet.cell(row=row_id, column=col_id).value = float(str(y).replace('[', '').replace(']', ''))
                    row_id += 1
                wb.save(tarfile)
                print('P2 response time has been updated to response sheet')
            except Exception as e:
                print('error occurred when update P2 data to response time sheet in Kun report', e)
        else:
            print('file of kun report is opened by another program')

