import os
import pandas as pd
import openpyxl
import numpy as np
from .incident_status_optimized import sheet1_update


"""""""""""""""""""""""""""
author: Shengzhen.Fu
purpose: load ActionItemsList then update Open action sheet in PMO report
version: 1.0.0
date: 2/Feb/2019
"""""""""""""""""""""""""""


class sheet3_update:

    def read_action_data(srcfile):
        print('start to read open action data from action item list file')
        if os.path.exists(srcfile):

            df = pd.read_excel(srcfile, sheet_name='Action Items List', usecols="F, H, L")
            px = df[(df['action items resolved']).isnull()].groupby(['team']).size().reset_index(
                name='Count of Action Items Ticket')
            py = px.sort_values(by=['Count of Action Items Ticket'], ascending=False)
            res_dict = py.to_dict('list')
            return res_dict
            print('complete read open action data from action item list file')
        else:
            print('file %s not found %s', srcfile)

    def update_open_action_data(srcfile, tarfile):
        try:
            print('loading kun report file')
            wb = openpyxl.load_workbook(filename=tarfile)
        except Exception as e:
            print('file of kun report is opened by another program', e)
        sheet = wb['Open Action Item']  # sheet name is Open Action Item
        row_id = 2
        res_dict = sheet3_update.read_action_data(srcfile)
        j = 0
        if len(res_dict.values()) > 0:
            print('begin to update open actionItems data to kun report')
            for x, y in enumerate(res_dict.values()):
                col = format(y).replace('[', '').replace(']', '').replace('\'', '')
                list_col = col.split(", ")
                for i in list_col:
                    if not i.isdigit():
                        sheet.cell(row=row_id, column=1).value = i
                        row_id += 1
                        j += 1
                    if i.isdigit():
                        sheet.cell(row=row_id - j, column=2).value = float(i)
                        row_id = row_id+1
                wb.save(tarfile)
            print('complete update open actionItems data to kun report')
        else:
            print('found no data in ActionItemsList201801_2018012, please check the data in file')

    def update_open_incident(srcfile, tarfile):
        """
        count open incident category by <20 days, 20 ~ 40 days, 41 ~ 60 days,  >60 days
        :param srcfile:
        :param tarfile:
        :return:
        """
        if os.path.exists(srcfile):
            # A key, H action items resolved I Duration
            pd.options.mode.chained_assignment = None
            df = pd.read_excel(srcfile, sheet_name='Action Items List', usecols="A,H,I,L")
            df1 = df[(df['action items resolved']).isnull()].groupby(['key'])['Duration'].max().round(0).reset_index(name='duration')
            df1['Duration'] = np.where(df1['duration'] > 60, '>60 Days', np.where(df1['duration'] < 20, '<20 Days',  np.where(df1['duration'] < 41, '20-40 Days', '41-60 Days')))
            df2 = pd.DataFrame(df1, columns=['Duration'])
            df3 = df2.groupby(['Duration']).size().to_frame('Incident Count').reset_index()
            print(df3)
            sheet1_update.append_df_to_excel(tarfile, df3, sheet_name='Open Action Item', header=None, index=False, startrow=1)

    def update_open_actionitem(srcfile, tarfile):
        """
        count open action items group by team category by <20 days, 20 ~ 40 days, 41 ~ 60 days,  >60 days
        :param srcfile:
        :param tarfile:
        :return:
        """
        if os.path.exists(srcfile):
            # column F action items ticket, H action items resolved I Duration L team
            pd.options.mode.chained_assignment = None
            df = pd.read_excel(srcfile, sheet_name='Action Items List', usecols="F,H,I,L")
            df1 = df[(df['action items resolved']).isnull()]
            df1['duration'] = np.where(df1['Duration'] > 60, '>60 Days', np.where(df1['Duration'] < 20, '<20 Days',  np.where(df1['Duration'] < 41, '20-40 Days', '41-60 Days')))
            df2 = pd.DataFrame(df1, columns=['duration', 'team'])
            df3 = df2.groupby(['duration', 'team']).size().to_frame('AI Count').reset_index()
            print(df3)
            sheet1_update.append_df_to_excel(tarfile, df3, sheet_name='Open Action Item', header=None, index=False, startrow=10)

    def update_open_action_data_new(srcfile, tarfile):
        """
        combine above 2 updates into 1
        :param srcfile:
        :param tarfile:
        :return:
        """
        sheet3_update.update_open_incident(srcfile, tarfile)
        sheet3_update.update_open_actionitem(srcfile, tarfile)
