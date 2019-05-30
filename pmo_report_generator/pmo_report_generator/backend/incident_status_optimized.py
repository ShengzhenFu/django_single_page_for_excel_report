import pandas as pd
import os
import time
import datetime
import openpyxl
from threading import Thread
import numpy as np


"""""""""""""""""""""""""""
author: Shengzhen.Fu
purpose: load created & resolved actionItems data from ActionItemsList file then update Incident Status in kun report
version: 1.0.0
date: 2/Feb/2019
"""""""""""""""""""""""""""


class sheet1_update:

    def get_last_day_previous_month():
        from datetime import date
        from dateutil.relativedelta import relativedelta
        today = date.today()
        return date(today.year, today.month, 1) - relativedelta(days=1)

    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        import pandas as pd

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

    def update_backlog(srcfile, tarfile):
        """
        update the open action items backlog category by <20 days, 20 ~ 40 days, 41 ~ 60 days,  >60 days
        """
        if os.path.exists(srcfile):
            # column A key, B priority C summary H action items resolved I Duration L team
            df = pd.read_excel(srcfile, sheet_name='Action Items List', usecols="A,B,C,H,I,L")
            df1 = df[(df['action items resolved']).isnull()].groupby(['key', 'priority', 'summary', 'team'])['Duration'].max().round(0).reset_index(name='duration')
            # df1['open duration'] = np.where(df1['duration'] > 60, '>60 Days', np.where(df1['duration'] < 20, '<20 Days', '20-40 Days'))
            df1['open duration'] = np.where(df1['duration'] > 60, '>60 Days', np.where(df1['duration'] < 20, '<20 Days', np.where(df1['duration'] < 41, '20-40 Days', '41-60 Days')))
            df2 = pd.DataFrame(df1, columns=['key', 'priority', 'summary', 'open duration', 'team'])
            print(df2)
            sheet1_update.append_df_to_excel(tarfile, df2, sheet_name='Incident Status', index=False, startrow=23)

    def incident_created_rca_key(month, file):
        """
        count incident created for one specified month
        """
        month_obj = datetime.datetime.strptime(str(month), "%Y%m")
        i_month = month_obj.strftime("%Y-%m")
        create_list = {}
        if os.path.exists(file):
            # column D created time column E rca key F action items ticket H action items resolved
            df = pd.read_excel(file, sheet_name='Action Items List', usecols="D,E,F,H")
            df.loc[:, ('yearMonth')] = df['created'].map(lambda x: 100 * x.year + x.month)
            m = df[(df['created'] >= '2018-01-01') & (df['yearMonth'].astype(str)==month)]
            # print(df.dtypes)
            # print(m['RCA Key'])
            n = set(m['rca key'].tolist())
            create_list.update({i_month: len(n)})
            # print('incident_created_rca_key called and returned ', create_list)
            return create_list
        else:
            print('file %s not found' % file)
            return None

    def update_month_incident(mon_list, tarfile):
        """
        update month column
        """
        t0 = time.time()
        try:
            wb = openpyxl.load_workbook(filename=tarfile)
            sheet = wb['Incident Status']  # sheet name is Incident Status
            row_id = 3
            col_id = 1
            for x in mon_list:  # insert month (eg,2018-01) to column 1, starting from row 3
                sheet.cell(row=row_id, column=col_id).value = str(x).replace('[', '').replace(']', '').replace('\'', '')
                row_id += 1
            wb.save(tarfile)
            print('month has been updated to incident status sheet')
            t1 = time.time()
            print('took %s seconds for month list update', t1 - t0)
        except Exception as e:
            print('error occurred when update month list to incident status sheet in Kun report', e)

    def update_created_incident(c_list, tarfile):
        """
        update created incident column
        """
        t0 = time.time()
        try:
            wb = openpyxl.load_workbook(filename=tarfile)
            sheet = wb['Incident Status']  # sheet name is Incident Status
            row_id = 3
            col_id = 2
            for y in c_list:  # insert 'Created' number to column 2, starting from row 3
                sheet.cell(row=row_id, column=col_id).value = int(str(y).replace('[', '').replace(']', ''))
                row_id += 1
            wb.save(tarfile)
            print('created num has been updated to incident status sheet')
            t1 = time.time()
            print('took %s seconds for created num update', t1 - t0)
        except Exception as e:
            print('error occurred when update created num to incident status sheet in Kun report', e)

    def update_resolved_incident(r_list, tarfile):
        """
        update resolved incident column
        """
        t0 = time.time()
        try:
            wb = openpyxl.load_workbook(filename=tarfile)
            sheet = wb['Incident Status']  # sheet name is Incident Status
            row_id = 3
            col_id = 3
            for z in r_list:  # insert 'resolved' num to column 3, starting from row 3
                sheet.cell(row=row_id, column=col_id).value = int(str(z).replace('[', '').replace(']', ''))
                row_id += 1
            wb.save(tarfile)
            print('resolved num has been updated to incident status sheet \n ')
            t1 = time.time()
            print('took %s seconds for resolved num update', t1 - t0)
        except Exception as e:
            print('error occurred when update resolved num to incident status sheet in Kun report', e)

    def resolved_rca_count_per_month(src_action_file):
        """
        get resolved incident data per monthly
        :param src_action_file:
        :return:
        """

        df_open = pd.read_excel(src_action_file, sheet_name='Action Items List', usecols="E,H")
        """""""""""""""""""""""""""""""""""""""""""""""""""
        find out the rca key that has action items not resolved
        """""""""""""""""""""""""""""""""""""""""""""""""""
        df_open = df_open[(df_open['action items resolved'].isnull())].groupby(['rca key'])['action items resolved'].max().reset_index(name='resolved')
        df_open.loc[:, ('key')] = df_open['rca key']
        df_open = pd.DataFrame(df_open, columns=['key'])
        list_rca = list(df_open.key)
        # print('rca key that has not resolved action items \n', list_rca)

        pd.set_option('display.max_rows', 500)
        pd.set_option('display.max_columns', 10)
        df = pd.read_excel(src_action_file, sheet_name='Action Items List', usecols="D,E,F,H")
        """""""""""""""""""""""""""""""""""""""""""""""""""
        find rca key that don't have action items NOT resolved
        """""""""""""""""""""""""""""""""""""""""""""""""""
        df = df[(df['created'] >= '2018-01-01') & (df['action items resolved'].notnull()) & (df['rca key'].isin(list_rca) == False)]

        df.loc[:, ('Resolved_Month')] = df['action items resolved'].map(lambda x: 100 * x.year + x.month)

        df1 = pd.DataFrame(df, columns=['Resolved_Month', 'rca key'])
        df1 = df1.sort_values(by=['Resolved_Month'])

        """""""""""""""""""""""""""""""""""""""""""""""""""
        for thsoe RCA has multiple action items, keep the last resolved one as the RCA resolved month 
        """""""""""""""""""""""""""""""""""""""""""""""""""
        df1 = df1.drop_duplicates(['rca key'], keep='last')
        """""""""""""""""""""""""""""""""""""""""""""""""""
        count rca key group by Resolved_Month 
        """""""""""""""""""""""""""""""""""""""""""""""""""
        m = df1.groupby(['Resolved_Month']).size().reset_index(name='Resolved')
        """""""""""""""""""""""""""""""""""""""""""""""""""
        change dtype of Resolved_Month to string 
        """""""""""""""""""""""""""""""""""""""""""""""""""
        n = m.astype({"Resolved_Month": str})
        """""""""""""""""""""""""""""""""""""""""""""""""""
        add column Month, string format YYYY-MM
        """""""""""""""""""""""""""""""""""""""""""""""""""
        n.loc[:, ('Month')] = n['Resolved_Month'].str[:4] + '-' + n['Resolved_Month'].str[4:6]
        """""""""""""""""""""""""""""""""""""""""""""""""""
        choose Month, Resolved column
        """""""""""""""""""""""""""""""""""""""""""""""""""
        df2 = pd.DataFrame(n, columns=['Month', 'Resolved']).sort_values(by=['Month'])

        dict_df2 = df2.to_dict('records')

        res_list = {}
        for k, v in enumerate(dict_df2):
            res_list.update({v['Month']: v['Resolved']})
        return res_list

    def update_created_resolved_incident(srcfile, tarfile):
        """
        read data from Action Items List 201801_201905.xlsx [Action Items List] sheet
        generate month list
        calculate RCA created in each month (count rca key by 'created' by month)
        calculate RCA resolved in each month (count rca key by 'action items resolved' by month, if there are multiple action items
            in one rca, get the latest resolved month)
        """
        print('start to update created & resolved incidents data')
        t0 = time.time()
        start_date = "2018-01-01"  # input start date
        end_date = sheet1_update.get_last_day_previous_month()  # datetime.datetime.now()  # input end date
        month_list = [i.strftime("%Y%m") for i in pd.date_range(start=start_date, end=end_date, freq='MS')]
        p = {}
        c_list, mon_list, r_list = [], [], []
        t1 = time.time()
        print('took %s seconds for month_list update', t1 - t0)
        for i_month in month_list:
            p.update(sheet1_update.incident_created_rca_key(i_month, srcfile))
        # print(p)
        t2 = time.time()
        print('took %s seconds for p update', t2 - t1)
        for key, value in p.items():
            mon = [key]
            mon_list.append(mon)
        print(mon_list)
        t3 = time.time()
        print('took %s seconds for mon_list update', t3 - t2)
        for key, value in p.items():
            c_num = [value]
            c_list.append(c_num)
        t4 = time.time()
        print('took %s seconds for c_list update', t4 - t3)

        res_list = sheet1_update.resolved_rca_count_per_month(srcfile)
        s_mon_list = []
        for s_month in mon_list:
            s_mon_list.append(str(s_month).replace('[', '').replace(']', '').replace('\'', ''))
        # print(s_mon_list)
        for i in s_mon_list:
            if i in res_list.keys():
                r_list.append(res_list[i])
            elif i not in res_list.keys():
                r_list.append(0)
        # print(r_list)
        t5 = time.time()
        print('took %s seconds for r_list update', t5 - t4)
        if os.path.exists(tarfile):
            # multiThread do update
            p1 = Thread(target=sheet1_update.update_month_incident(mon_list, tarfile))
            p2 = Thread(target=sheet1_update.update_created_incident(c_list, tarfile))
            p3 = Thread(target=sheet1_update.update_resolved_incident(r_list, tarfile))
            p1.start()
            p2.start()
            p3.start()
            t6 = time.time()
            print('total took %s seconds for incident status update', t6 - t0)
        else:
            print('target file not exist, please check')
        sheet1_update.update_backlog(srcfile, tarfile)
